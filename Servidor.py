from flask import Flask, request, jsonify, render_template, send_file
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import os

app = Flask(__name__)

EXCEL_FILE = "registros_impala.xlsx"

CHECKLIST_ITEMS = [
    "extintores_ok",
    "calcomania_seguro_ok",
    "llantas_ok",
    "viene_vacia",
    "luces_ok",
    "kit_emergencia_ok",
    "documentos_conductor_ok",
    "mangueras_valvulas_ok",
    "sin_fugas_ok",
    "equipo_proteccion_ok",
]

CHECKLIST_LABELS = {
    "extintores_ok": "Extintores en buen estado",
    "calcomania_seguro_ok": "Calcomanía del seguro vigente",
    "llantas_ok": "Llantas en buen estado",
    "viene_vacia": "Unidad viene vacía",
    "luces_ok": "Luces funcionando",
    "kit_emergencia_ok": "Kit de emergencia/derrame",
    "documentos_conductor_ok": "Documentos del conductor",
    "mangueras_valvulas_ok": "Mangueras y válvulas OK",
    "sin_fugas_ok": "Sin fugas visibles",
    "equipo_proteccion_ok": "Equipo de protección personal",
}

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Registros"

        header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        headers = ["Fecha", "Hora", "ID Unidad", "Placas", "Conductor"] + \
                  [CHECKLIST_LABELS[k] for k in CHECKLIST_ITEMS] + \
                  ["Resultado", "Observaciones"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20
        for i in range(6, 6 + len(CHECKLIST_ITEMS)):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(6 + len(CHECKLIST_ITEMS))].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(7 + len(CHECKLIST_ITEMS))].width = 30

        ws.row_dimensions[1].height = 40
        wb.save(EXCEL_FILE)

def save_to_excel(data):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Registros"]

    now = datetime.now()
    checklist_values = [data.get(k, "NO") for k in CHECKLIST_ITEMS]
    failed = [CHECKLIST_LABELS[k] for k in CHECKLIST_ITEMS if data.get(k) != "SI"]
    resultado = "APROBADO" if not failed else "RECHAZADO"

    row_data = [
        now.strftime("%d/%m/%Y"),
        now.strftime("%H:%M:%S"),
        data.get("id_unidad", ""),
        data.get("placas", ""),
        data.get("conductor", ""),
    ] + checklist_values + [resultado, data.get("observaciones", "")]

    next_row = ws.max_row + 1
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    green_fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
    red_fill = PatternFill(start_color="f8d7da", end_color="f8d7da", fill_type="solid")
    row_fill = PatternFill(start_color="f9f9f9", end_color="f9f9f9", fill_type="solid") if next_row % 2 == 0 else None

    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=next_row, column=col, value=value)
        cell.alignment = center
        cell.border = thin
        if row_fill:
            cell.fill = row_fill

    # Color checklist cells
    for i, key in enumerate(CHECKLIST_ITEMS):
        col = 6 + i
        cell = ws.cell(row=next_row, column=col)
        val = data.get(key, "NO")
        cell.fill = green_fill if val == "SI" else red_fill
        cell.font = Font(bold=True, color="155724" if val == "SI" else "721c24")

    # Color resultado cell
    res_col = 6 + len(CHECKLIST_ITEMS)
    res_cell = ws.cell(row=next_row, column=res_col)
    res_cell.fill = green_fill if resultado == "APROBADO" else red_fill
    res_cell.font = Font(bold=True, color="155724" if resultado == "APROBADO" else "721c24")

    wb.save(EXCEL_FILE)
    return resultado, failed

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/registrar", methods=["POST"])
def registrar():
    data = request.json
    resultado, fallidos = save_to_excel(data)
    return jsonify({
        "success": True,
        "resultado": resultado,
        "fallidos": fallidos,
        "mensaje": "Registro guardado correctamente"
    })

@app.route("/exportar")
def exportar():
    init_excel()
    return send_file(EXCEL_FILE, as_attachment=True, download_name="registros_impala.xlsx")

@app.route("/registros")
def ver_registros():
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Registros"]
    registros = []
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            registros.append(dict(zip(headers, row)))
    return jsonify(registros)

if __name__ == "__main__":
    init_excel()
    print("=" * 50)
    print("  IMPALA TERMINALS - Sistema de Inspección")
    print("=" * 50)
    print(f"  Servidor iniciado.")
    print(f"  Abre en la tablet: http://<IP-DE-ESTA-PC>:5000")
    print(f"  Excel se guarda en: {os.path.abspath(EXCEL_FILE)}")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000, debug=False)
