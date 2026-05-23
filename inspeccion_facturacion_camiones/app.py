from __future__ import annotations

from datetime import datetime
import csv
import io
import json
import os
from pathlib import Path
import sys

BASE_DIR = Path(__file__).resolve().parent
LOCAL_PACKAGES_DIR = BASE_DIR / ".python_packages"
try:
    local_flask_init = LOCAL_PACKAGES_DIR / "flask" / "__init__.py"
    if local_flask_init.exists():
        sys.path.insert(0, str(LOCAL_PACKAGES_DIR))
except OSError:
    pass

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from flask import Flask, flash, g, jsonify, redirect, render_template, request, send_file, url_for

app = Flask(__name__)
app.config["SECRET_KEY"] = "cambia-esta-clave-en-produccion"

DB_CONFIG = {
    "host": os.getenv("IMPALA_DB_HOST", "localhost"),
    "user": os.getenv("IMPALA_DB_USER", "impala_app"),
    "password": os.getenv("IMPALA_DB_PASSWORD", ""),
    "database": os.getenv("IMPALA_DB_NAME", "inspeccion_camiones"),
    "port": int(os.getenv("IMPALA_DB_PORT", "3306")),
    "cursorclass": pymysql.cursors.DictCursor,
    "autocommit": False,
}

CHECKLIST_SECTIONS = [
    {
        "code": "A",
        "title": "Cabezal serie A",
        "description": "Revision del cabezal y condiciones de ingreso de la unidad tractora.",
        "fields": [
            {"name": "a1_bateria", "label": "A.1 Bateria: sujecion, aislamiento y proteccion"},
            {"name": "a2_master_switch", "label": 'A.2 Interruptor maestro ("Master Switch")'},
            {"name": "a3_cableado_sellado", "label": "A.3 Cableado y sistema electrico sellado"},
            {"name": "a4_union_cables", "label": "A.4 Union de cables electricos con cinta vulcanizada y corrugado"},
            {"name": "a5_luces", "label": "A.5 Luces sin conexiones expuestas ni tapaderas faltantes"},
            {"name": "a6_escape_supresor", "label": "A.6 Sistema de escape con tuberia en buen estado y supresor de flama"},
            {"name": "a7_escape_fugas", "label": "A.7 Sistema de escape del camion rigido protegido contra fugas"},
            {"name": "a8_escape_valvulas", "label": "A.8 Sistema de escape separado de las valvulas de descarga"},
            {"name": "a9_vidrio_plumillas", "label": "A.9 Vidrio delantero y plumillas en buen estado"},
            {"name": "a10_ventanas", "label": "A.10 Ventanas laterales cerradas durante ingreso y carga"},
            {"name": "a11_ventilacion", "label": "A.11 Apertura de ventilacion superior en cabina cerrada"},
            {"name": "a12_extintor", "label": "A.12 Extintor ABC de 10 lb vigente y de facil acceso"},
            {"name": "a13_tanques_fugas", "label": "A.13 Tanques de diesel, aceite y sistema de aire sin fugas"},
            {"name": "a14_neumaticos", "label": "A.14 Neumaticos en buen estado"},
            {"name": "a15_retroceso", "label": "A.15 Alarma y luz de retroceso en funcionamiento"},
        ],
    },
    {
        "code": "B",
        "title": "Cisterna serie B",
        "description": "Validacion documental, mecanica y de seguridad de la cisterna.",
        "fields": [
            {"name": "b1_tabla_calibracion", "label": "B.1 Tabla de calibracion vigente", "note_placeholder": "Anota fecha de vencimiento si aplica"},
            {"name": "b2_licencia_mem", "label": "B.2 Licencia MEM de la unidad vigente", "note_placeholder": "Anota fecha de vencimiento si aplica"},
            {"name": "b3_seguro_unidad", "label": "B.3 Seguro de la unidad y calcomania vigente", "note_placeholder": "Anota fecha de vencimiento si aplica"},
            {"name": "b4_cableado_sellado", "label": "B.4 Cableado y sistema electrico sellado"},
            {"name": "b5_union_cables", "label": "B.5 Union de cables electricos con cinta vulcanizada y corrugado"},
            {"name": "b6_luces", "label": "B.6 Luces sin conexiones expuestas ni rajaduras"},
            {"name": "b7_neumaticos", "label": "B.7 Neumaticos en buen estado"},
            {"name": "b8_extintor", "label": "B.8 Extintor ABC de 20 lb vigente y de facil acceso"},
            {"name": "b9_compartimientos", "label": "B.9 Compartimientos vacios para evitar sobrellenado o contaminacion"},
            {"name": "b10_rotulacion", "label": "B.10 Rotulacion de compartimientos y tabla de calibracion"},
            {"name": "b11_venteo", "label": "B.11 Valvula de venteo o desfogue por compartimiento"},
            {"name": "b12_fondo_seguridad", "label": "B.12 Valvula de fondo de seguridad por compartimiento"},
            {"name": "b13_tierra", "label": "B.13 Puntos de conexion a tierra sin pintura"},
            {"name": "b14_cuerpo_fugas", "label": "B.14 Cuerpo del tanque, tuberias y valvulas sin manchas de fugas"},
            {"name": "b15_descarga_asegurada", "label": "B.15 Valvulas de descarga aseguradas para marchamos y cierre rapido"},
            {"name": "b16_continuidad", "label": "B.16 Continuidad de compartimiento con cable o tubo"},
            {"name": "b17_retroceso", "label": "B.17 Alarma y luz de retroceso en buen estado"},
            {"name": "b18_rotulacion_cisterna", "label": "B.18 Rotulacion de cisterna con cinta reflectiva y rombo de identificacion"},
            {"name": "b19_manholes", "label": 'B.19 "Manholes" con altura maxima de 3 pulgadas y bajo proteccion de volcadura'},
            {"name": "b20_barra_valvulas", "label": "B.20 Barra de proteccion de valvulas"},
            {"name": "b21_vacio_compartimientos", "label": 'B.21 Vacio minimo de 5 pulgadas desde la base de los "manholes"'},
        ],
    },
    {
        "code": "C",
        "title": "Conductor serie C",
        "description": "Revision del conductor, controles de acceso y requisitos de seguridad.",
        "fields": [
            {"name": "c1_licencia_profesional", "label": "C.1 Licencia profesional vigente"},
            {"name": "c2_alcoholemia", "label": "C.2 Prueba de alcoholemia en 0 grados", "note_placeholder": "Anota porcentaje o lectura obtenida"},
            {"name": "c3_epp", "label": "C.3 Equipo de proteccion personal completo", "note_placeholder": "Indica EPP faltante si aplica"},
            {"name": "c4_armas", "label": "C.4 Prohibicion de ingreso de armas de fuego"},
            {"name": "c5_celulares", "label": "C.5 Restriccion de ingreso de celulares o dispositivos electronicos"},
            {"name": "c6_entrenamiento", "label": "C.6 Certificacion vigente de entrenamiento de carga en terminal", "note_placeholder": "Anota fecha de vencimiento"},
        ],
    },
]

CHECKLIST_FIELDS = [
    {**field, "section_code": section["code"], "section_title": section["title"]}
    for section in CHECKLIST_SECTIONS
    for field in section["fields"]
]

STATUS_LABELS = {
    "ok": "Cumple",
    "fail": "No cumple",
}

OPERATION_OPTIONS = ["Ex-Rack", "Flota propia"]
PRODUCT_OPTIONS = ["Limpios", "Fuel Oíl"]
RESULT_OPTIONS = ["Cumple", "No cumple"]
FALLBACK_CHECKLIST_BY_RESULT = {
    "Cumple": "Cumple",
    "No cumple": "No cumple",
    "Aprobado": "Cumple",
    "Condicional": "Cumple",
    "Rechazado": "No cumple",
}

LEGACY_RESULT_MAP = {
    "Aprobado": "Cumple",
    "Condicional": "Cumple",
    "Rechazado": "No cumple",
}

LEGACY_CHECKLIST_STATUS_MAP = {
    "warning": "ok",
}


def get_db():
    if "db" not in g:
        g.db = pymysql.connect(**DB_CONFIG)
    return g.db


@app.teardown_appcontext
def close_db(_error: Exception | None) -> None:
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS inspecciones (
                id INT NOT NULL AUTO_INCREMENT PRIMARY KEY,
                created_at DATETIME NOT NULL,
                guardia_turno VARCHAR(150) NOT NULL,
                hora_entrada VARCHAR(10) NOT NULL,
                hora_salida VARCHAR(10),
                tipo_operacion VARCHAR(30) NOT NULL,
                numero_unidad VARCHAR(100) NOT NULL,
                placa VARCHAR(100) NOT NULL,
                piloto VARCHAR(150) NOT NULL,
                empresa VARCHAR(150) NOT NULL,
                producto VARCHAR(150) NOT NULL,
                documento_referencia VARCHAR(150),
                inspector_salida VARCHAR(150),
                cantidad_marchamos INT,
                correlativos_marchamos TEXT,
                observaciones_salida TEXT,
                inspector VARCHAR(150) NOT NULL,
                resultado_final VARCHAR(50) NOT NULL,
                checklist_json LONGTEXT NOT NULL,
                observaciones_generales TEXT
            )
            """
        )
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS guardia_turno VARCHAR(150) NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS hora_entrada VARCHAR(10) NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS hora_salida VARCHAR(10) NULL")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS tipo_operacion VARCHAR(30) NOT NULL DEFAULT ''")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS inspector_salida VARCHAR(150) NULL")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS cantidad_marchamos INT NULL")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS correlativos_marchamos TEXT NULL")
        cursor.execute("ALTER TABLE inspecciones ADD COLUMN IF NOT EXISTS observaciones_salida TEXT NULL")
    db.commit()


def parse_checklist(form_data: dict[str, str]) -> tuple[list[dict[str, str]], list[str]]:
    checklist: list[dict[str, str]] = []
    errors: list[str] = []

    for field in CHECKLIST_FIELDS:
        status = form_data.get(field["name"], "").strip()
        if status not in STATUS_LABELS:
            errors.append(f"Debes indicar el estado de: {field['label']}.")
            continue
        checklist.append(
            {
                "name": field["name"],
                "label": field["label"],
                "section_code": field["section_code"],
                "section_title": field["section_title"],
                "status": status,
                "status_label": STATUS_LABELS[status],
            }
        )

    return checklist, errors


def validate_form(form_data: dict[str, str]) -> tuple[dict[str, str], list[dict[str, str]], list[str]]:
    payload = {
        "guardia_turno": form_data.get("guardia_turno", "").strip(),
        "hora_entrada": form_data.get("hora_entrada", "").strip(),
        "tipo_operacion": form_data.get("tipo_operacion", "").strip(),
        "numero_unidad": form_data.get("numero_unidad", "").strip(),
        "placa": form_data.get("placa", "").strip(),
        "piloto": form_data.get("piloto", "").strip(),
        "empresa": form_data.get("empresa", "").strip(),
        "producto": form_data.get("producto", "").strip(),
        "documento_referencia": form_data.get("documento_referencia", "").strip(),
        "inspector": form_data.get("guardia_turno", "").strip(),
        "resultado_final": form_data.get("resultado_final", "").strip(),
        "observaciones_generales": form_data.get("observaciones_generales", "").strip(),
    }

    required_fields = {
        "guardia_turno": "Guardia de turno",
        "hora_entrada": "Hora de entrada",
        "tipo_operacion": "Tipo de operacion",
        "numero_unidad": "Numero de unidad",
        "placa": "Placa / TC",
        "piloto": "Nombre del piloto",
        "empresa": "Empresa transportista",
        "producto": "Tipo de producto",
        "documento_referencia": "Numero de pedido",
        "resultado_final": "Resultado final",
    }

    errors = [
        f"El campo '{label}' es obligatorio."
        for key, label in required_fields.items()
        if not payload[key]
    ]

    if payload["resultado_final"] and payload["resultado_final"] not in RESULT_OPTIONS:
        errors.append("El resultado final no es valido.")

    if payload["producto"] and payload["producto"] not in PRODUCT_OPTIONS:
        errors.append("El tipo de producto no es valido.")

    if payload["tipo_operacion"] and payload["tipo_operacion"] not in OPERATION_OPTIONS:
        errors.append("El tipo de operacion no es valido.")

    checklist, checklist_errors = parse_checklist(form_data)
    errors.extend(checklist_errors)

    return payload, checklist, errors


def validate_exit_form(form_data: dict[str, str]) -> tuple[dict[str, str], list[str]]:
    exit_data = {
        "documento_referencia": form_data.get("documento_referencia_salida", "").strip(),
        "hora_salida": form_data.get("hora_salida", "").strip(),
        "inspector_salida": form_data.get("inspector_salida", "").strip(),
        "cantidad_marchamos": form_data.get("cantidad_marchamos", "").strip(),
        "correlativos_marchamos": form_data.get("correlativos_marchamos", "").strip(),
        "observaciones_salida": form_data.get("observaciones_salida", "").strip(),
    }

    required_fields = {
        "documento_referencia": "Numero de pedido",
        "hora_salida": "Hora de salida",
        "inspector_salida": "Inspector de salida",
        "cantidad_marchamos": "Cantidad de marchamos",
        "correlativos_marchamos": "Correlativos de marchamos",
    }

    errors = [
        f"El campo '{label}' es obligatorio."
        for key, label in required_fields.items()
        if not exit_data[key]
    ]

    if exit_data["cantidad_marchamos"]:
        try:
            cantidad = int(exit_data["cantidad_marchamos"])
            if cantidad < 0:
                errors.append("La cantidad de marchamos no puede ser negativa.")
        except ValueError:
            errors.append("La cantidad de marchamos debe ser numerica.")

    return exit_data, errors


def create_inspection(payload: dict[str, str], checklist: list[dict[str, str]]) -> int:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO inspecciones (
                created_at,
                guardia_turno,
                hora_entrada,
                hora_salida,
                tipo_operacion,
                numero_unidad,
                placa,
                piloto,
                empresa,
                producto,
                documento_referencia,
                inspector_salida,
                cantidad_marchamos,
                correlativos_marchamos,
                observaciones_salida,
                inspector,
                resultado_final,
                checklist_json,
                observaciones_generales
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                datetime.now(),
                payload["guardia_turno"],
                payload["hora_entrada"],
                None,
                payload["tipo_operacion"],
                payload["numero_unidad"],
                payload["placa"],
                payload["piloto"],
                payload["empresa"],
                payload["producto"],
                payload["documento_referencia"],
                None,
                None,
                None,
                None,
                payload["inspector"],
                payload["resultado_final"],
                json.dumps(checklist, ensure_ascii=False),
                payload["observaciones_generales"],
            ),
        )
        new_id = int(cursor.lastrowid)
    db.commit()
    return new_id


def update_exit_by_order(exit_data: dict[str, str]) -> bool:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            UPDATE inspecciones
            SET
                hora_salida = %s,
                inspector_salida = %s,
                cantidad_marchamos = %s,
                correlativos_marchamos = %s,
                observaciones_salida = %s
            WHERE documento_referencia = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (
                exit_data["hora_salida"],
                exit_data["inspector_salida"],
                int(exit_data["cantidad_marchamos"]),
                exit_data["correlativos_marchamos"],
                exit_data["observaciones_salida"] or None,
                exit_data["documento_referencia"],
            ),
        )
        updated = cursor.rowcount > 0
    db.commit()
    return updated


def normalize_result_label(value: str | None) -> str:
    if not value:
        return ""
    return LEGACY_RESULT_MAP.get(value, value)


def normalize_checklist_entries(checklist: list[dict]) -> list[dict]:
    normalized: list[dict] = []
    for check in checklist:
        item = dict(check)
        status = item.get("status", "")
        status = LEGACY_CHECKLIST_STATUS_MAP.get(status, status)
        item["status"] = status
        if status in STATUS_LABELS:
            item["status_label"] = STATUS_LABELS[status]
        normalized.append(item)
    return normalized


def fetch_recent_inspections(limit: int = 20) -> list[dict]:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT *
            FROM inspecciones
            ORDER BY id DESC
            LIMIT %s
            """,
            (limit,),
        )
        rows = cursor.fetchall()

    inspections: list[dict] = []
    for row in rows:
        item = dict(row)
        if isinstance(item.get("created_at"), datetime):
            item["created_at"] = item["created_at"].strftime("%Y-%m-%d %H:%M:%S")
        item["resultado_final"] = normalize_result_label(item.get("resultado_final"))
        item["checklist"] = normalize_checklist_entries(json.loads(item["checklist_json"] or "[]"))
        inspections.append(item)
    return inspections


def fetch_inspection_by_order(order_number: str) -> dict | None:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            """
            SELECT *
            FROM inspecciones
            WHERE documento_referencia = %s
            ORDER BY id DESC
            LIMIT 1
            """,
            (order_number,),
        )
        row = cursor.fetchone()

    if row is None:
        return None

    item = dict(row)
    if isinstance(item.get("created_at"), datetime):
        item["created_at"] = item["created_at"].strftime("%Y-%m-%d %H:%M:%S")
    item["resultado_final"] = normalize_result_label(item.get("resultado_final"))
    item["checklist"] = normalize_checklist_entries(json.loads(item["checklist_json"] or "[]"))
    return item


def fetch_dashboard_stats() -> dict[str, int]:
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute("SELECT COUNT(*) AS total FROM inspecciones WHERE DATE(created_at) = CURDATE()")
        total_today = cursor.fetchone()["total"]
        cursor.execute(
            "SELECT COUNT(*) AS total FROM inspecciones WHERE DATE(created_at) = CURDATE() AND resultado_final IN (%s, %s, %s)",
            ("Cumple", "Aprobado", "Condicional"),
        )
        approved = cursor.fetchone()["total"]
        cursor.execute(
            "SELECT COUNT(*) AS total FROM inspecciones WHERE DATE(created_at) = CURDATE() AND resultado_final IN (%s, %s)",
            ("No cumple", "Rechazado"),
        )
        rejected = cursor.fetchone()["total"]
    return {
        "total_today": total_today,
        "approved": approved,
        "rejected": rejected,
    }


def build_export_headers() -> list[str]:
    headers = [
        "ID",
        "Fecha",
        "Hora registro",
        "Guardia",
        "Hora entrada",
        "Hora salida",
        "Inspector salida",
        "Cantidad marchamos",
        "Correlativos marchamos",
        "Observaciones salida",
        "Tipo de operacion",
        "Unidad",
        "Placa",
        "Piloto",
        "Empresa",
        "Tipo de producto",
        "Numero de pedido",
        "Inspector ingreso",
        "Resultado final",
        "Nota de respaldo",
    ]

    for field in CHECKLIST_FIELDS:
        headers.append(field["label"])

    return headers


def flatten_inspection_for_export(item: dict) -> list[str]:
    checklist_by_name = {check.get("name"): check for check in item["checklist"] if check.get("name")}
    checklist_by_label = {check.get("label"): check for check in item["checklist"] if check.get("label")}

    created_at = item["created_at"]
    created_date = created_at
    created_time = ""
    if isinstance(created_at, str) and " " in created_at:
        created_date, created_time = created_at.split(" ", 1)

    row = [
        item["id"],
        created_date,
        created_time,
        item.get("guardia_turno", ""),
        item.get("hora_entrada", ""),
        item.get("hora_salida") or "",
        item.get("inspector_salida") or "",
        item.get("cantidad_marchamos") or "",
        item.get("correlativos_marchamos") or "",
        item.get("observaciones_salida") or "",
        item.get("tipo_operacion", ""),
        item["numero_unidad"],
        item["placa"],
        item["piloto"],
        item["empresa"],
        item["producto"],
        item["documento_referencia"],
        item["inspector"],
        normalize_result_label(item["resultado_final"]),
        item["observaciones_generales"],
    ]

    for field in CHECKLIST_FIELDS:
        check = checklist_by_name.get(field["name"]) or checklist_by_label.get(field["label"]) or {}
        status_label = check.get("status_label")
        if not status_label and check.get("status") in STATUS_LABELS:
            status_label = STATUS_LABELS[check["status"]]
        if not status_label:
            status_label = FALLBACK_CHECKLIST_BY_RESULT.get(item.get("resultado_final", ""), "")

        row.append(status_label)

    return row


def build_excel_workbook(inspections: list[dict]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inspecciones"

    headers = build_export_headers()
    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="903F98")
    header_font = Font(color="FFFFFF", bold=True)

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = header_fill
        cell.font = header_font

    for item in inspections:
        sheet.append(flatten_inspection_for_export(item))

    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value_length = len(str(cell.value)) if cell.value is not None else 0
            if value_length > max_length:
                max_length = value_length
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 14), 38)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return workbook


def render_home(
    form_data: dict[str, str] | None = None,
    salida_form_data: dict[str, str] | None = None,
    status_code: int = 200,
):
    return (
        render_template(
            "index.html",
            checklist_sections=CHECKLIST_SECTIONS,
            checklist_fields=CHECKLIST_FIELDS,
            product_options=PRODUCT_OPTIONS,
            result_options=RESULT_OPTIONS,
            operation_options=OPERATION_OPTIONS,
            inspections=fetch_recent_inspections(),
            stats=fetch_dashboard_stats(),
            status_labels=STATUS_LABELS,
            now=datetime.now().strftime("%d/%m/%Y"),
            form_data=form_data or {},
            salida_form_data=salida_form_data or {},
            epp_description="Casco, lentes transparentes, guantes de nitrilo, calzado con puntera, camisa manga larga u overol 100% algodon y arnes de cuerpo.",
        ),
        status_code,
    )


@app.before_request
def ensure_database() -> None:
    init_db()


@app.get("/")
def home():
    return render_home()


@app.post("/inspecciones")
def save_inspection():
    form_data = request.form.to_dict()
    payload, checklist, errors = validate_form(form_data)

    if errors:
        for error in errors:
            flash(error, "error")
        return render_home(form_data=form_data, status_code=400)

    new_id = create_inspection(payload, checklist)
    flash(f"Inspeccion #{new_id} guardada correctamente.", "success")
    return redirect(url_for("home"))


@app.post("/salidas")
def save_exit():
    form_data = request.form.to_dict()
    exit_data, errors = validate_exit_form(form_data)

    if errors:
        for error in errors:
            flash(error, "error")
        return render_home(salida_form_data=form_data, status_code=400)

    if not update_exit_by_order(exit_data):
        flash("No se encontro un ingreso registrado con ese numero de pedido.", "error")
        return render_home(salida_form_data=form_data, status_code=404)

    flash(f"Salida del pedido {exit_data['documento_referencia']} actualizada correctamente.", "success")
    return redirect(url_for("home"))


@app.get("/api/inspecciones")
def list_inspections_api():
    return jsonify(fetch_recent_inspections(limit=100))


@app.get("/api/inspecciones/buscar")
def search_inspection_by_order_api():
    order_number = request.args.get("pedido", "").strip()
    if not order_number:
        return jsonify({"error": "Debes indicar un numero de pedido."}), 400

    item = fetch_inspection_by_order(order_number)
    if item is None:
        return jsonify({"error": "Inspeccion no encontrada."}), 404

    return jsonify(item)


@app.get("/api/inspecciones/<int:inspection_id>")
def inspection_detail_api(inspection_id: int):
    db = get_db()
    with db.cursor() as cursor:
        cursor.execute(
            "SELECT * FROM inspecciones WHERE id = %s",
            (inspection_id,),
        )
        row = cursor.fetchone()
    if row is None:
        return jsonify({"error": "Inspeccion no encontrada."}), 404

    item = dict(row)
    if isinstance(item.get("created_at"), datetime):
        item["created_at"] = item["created_at"].strftime("%Y-%m-%d %H:%M:%S")
    item["resultado_final"] = normalize_result_label(item.get("resultado_final"))
    item["checklist"] = normalize_checklist_entries(json.loads(item["checklist_json"] or "[]"))
    return jsonify(item)


@app.get("/export/excel")
def export_excel():
    inspections = fetch_recent_inspections(limit=1000)
    workbook = build_excel_workbook(inspections)
    bytes_io = io.BytesIO()
    workbook.save(bytes_io)
    bytes_io.seek(0)
    filename = f"inspecciones_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        bytes_io,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    with app.app_context():
        init_db()
    app.run(host="0.0.0.0", port=5000, debug=True)
